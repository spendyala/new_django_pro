from openpyxl import Workbook
from openpyxl.styles import Style, Alignment, Font, PatternFill, Color, Fill

wb = Workbook()

ws = wb.active

# Tab Name
ws.title = "Steam Leak"

mycell = ws['A6']

# Cell Properties
cell_attributes = {'alignment': Alignment(horizontal='center'),
                   'font': Font(bold=True),
                   'fill': PatternFill(patternType='solid',
                                       fgColor=Color('90909090'))}
# Cell Style
cell_style = Style(**cell_attributes)
# Assigning the style to Cell
mycell.style = cell_style
# Assigning Text to Cell
ws['A6'] = "Steam Trap Assessment"
# Mergeing Cells
ws.merge_cells('A6:M6')

# Saving the Excel File
wb.save('sample_excel.xlsx')
