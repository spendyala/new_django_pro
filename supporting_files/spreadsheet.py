# import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Style, Alignment, Font, PatternFill, Color, Fill
import StringIO
from django.http import HttpResponse
from valve_insulation.models import INSULATION_CHOICES
from air_compressors.models import COMPRESSOR_TYPE, VFD_CONTROL_TYPE


class SteamLeakExcel(object):
    def __init__(self, steam_leak=None):
        self.steam_leak = steam_leak
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Leak #',
            'B1': 'Location/Description',
            'C1': 'Boiler Efficiency',
            'D1': 'Therms (Therm Rate)',
            'E1': 'Pressure (psig)',
            'F1': 'Absolute Pressure (psia)',
            'G1': 'Size Leak (inch)',
            'H1': 'Steam Loss (pph)',
            'I1': 'Steam Energy (Btu/lb)',
            'J1': 'Gas Usage (Therms/hour)',
            'K1': 'Cost / Hour',
            'L1': 'Hours of Operation',
            'M1': 'Cost / Year'
        }
        self.data_column = {
            'A': 'steam_leak_number',
            'B': 'location_description',
            'C': 'boiler_efficiency',
            'D': 'get_therm_rate',
            'E': 'pressure_in_psig',
            'F': 'absolute_pressure_psia',
            'G': 'size_leak_in_inch',
            'H': 'get_steam_loss_pph',
            'I': 'get_steam_energy_btu_per_lb',
            'J': 'get_gas_usage_therms_per_hour',
            'K': 'get_cost_per_hour',
            'L': 'hours_of_operation',
            'M': 'get_cost_per_year'
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        for each_steam_leak in self.steam_leak:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                self.ws[cell_val] = each_steam_leak.get(
                    self.data_column[each_column])
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'SteamLeak.xlsx')
        return response


class SteamTrapExcel(object):
    def __init__(self, steam_trap=None):
        self.steam_trap = steam_trap
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Failed Trap #',
            'B1': 'Location/Description',
            'C1': 'Boiler Efficiency',
            'D1': 'Therms (Therm Rate)',
            'E1': 'Pressure (psig)',
            'F1': 'Absolute Pressure (psia)',
            'G1': 'Trap Pipe Size (inch)',
            'H1': 'Size Trap (inch)',
            'I1': 'Steam Loss (pph)',
            'J1': 'Steam Energy (Btu/lb)',
            'K1': 'Gas Usage (Therms/hour)',
            'L1': 'Cost / Hour',
            'M1': 'Hours of Operation',
            'N1': 'Cost / Year'
        }
        self.data_column = {
            'A': 'steam_trap_number',
            'B': 'location_description',
            'C': 'boiler_efficiency',
            'D': 'get_therm_rate',
            'E': 'pressure_in_psig',
            'F': 'absolute_pressure_psia',
            'G': 'trap_pipe_size',
            'H': 'size_trap_orifice',
            'I': 'get_steam_loss_pph',
            'J': 'get_steam_energy_btu_per_lb',
            'K': 'get_gas_usage_therms_per_hour',
            'L': 'get_cost_per_hour',
            'M': 'hours_of_operation',
            'N': 'get_cost_per_year'
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        for each_steam_trap in self.steam_trap:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                self.ws[cell_val] = each_steam_trap.get(
                    self.data_column[each_column])
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'SteamTrap.xlsx')
        return response


class BoilerDatacollectionExcel(object):
    def __init__(self, boiler_datacollection=None):
        self.boiler_datacollection = boiler_datacollection
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Boiler Data Collection Name',
            'B1': 'Boiler capacity (MBH)',
            'C1': 'Yearly Hours of Operation',
            'D1': 'Separately Meter',
            'E1': 'Make Up Water Log/ Separate Bill',
            'F1': 'Number of Steam Traps',
            'G1': 'Date Last Steam Trap\nAudit Performed',
            'H1': 'Is the Header Insulated',
            'I1': 'Percent of Condensate\nthat returns to Boiler',
            'J1': 'Aerator Tank Temperature',
            'K1': 'Aerator Tank Pressure',
            'L1': 'Production Pressure',
        }
        self.data_column = {
            'A': 'name',
            'B': 'boiler_capacity_mbh',
            'C': 'hours_of_operation',
            'D': 'separately_meter',
            'E': 'make_up_water_log_separate_bill',
            'F': 'no_of_steam_traps',
            'G': 'steam_trap_audit_performed',
            'H': 'is_the_header_insulated',
            'I': 'percentage_condensate_that_returns_to_boiler',
            'J': 'aerator_tank_temp',
            'K': 'aerator_tank_pressure',
            'L': 'production_pressure',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        for each_boiler_datacollection in self.boiler_datacollection:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                if column_val in [
                        'separately_meter',
                        'make_up_water_log_separate_bill',
                        'is_the_header_insulated']:
                    self.ws[cell_val] = ('Yes' if
                                         each_boiler_datacollection.get(
                                            column_val)
                                         else 'No')
                else:
                    self.ws[cell_val] = each_boiler_datacollection.get(
                        column_val)
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'BoilerDatacollection.xlsx')
        return response


class StackedEconomizerExcel(object):
    def __init__(self, stacked_economizer=None):
        self.stacked_economizer = stacked_economizer
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Stacked Economizer',
            'B1': 'Hours of Operation',
            'C1': 'Boiler Size (HP)',
            'D1': 'Boiler size, MMBtu/hr',
            'E1': 'Initial Stack Gas Temp (F)',
            'F1': 'Average Fire Rate',
            'G1': 'Recoverable Heat, MMBtu/hr',
            'H1': 'Recoverable Heat, Therms/yr',
            'I1': 'Savings'
        }
        self.data_column = {
            'A': 'boiler_stacked_economizer',
            'B': 'hours_of_operations',
            'C': 'boiler_size_hp',
            'D': 'get_boiler_size_mmbtu_per_hr',
            'E': 'initial_stack_gas_temp_f',
            'F': 'average_fire_rate',
            'G': 'get_recoverable_heat_mmbtu_per_hr',
            'H': 'get_recoverable_heat_therms_per_year',
            'I': 'get_savings'
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        for each_stacked_economizer in self.stacked_economizer:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_stacked_economizer.get(column_val)
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'StackedEconomizer.xlsx')
        return response


class ValveInsulationExcel(object):
    def __init__(self, valve_insulation=None):
        self.valve_insulation = valve_insulation
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Valve Insulation Name',
            'B1': 'Valve Type',
            'C1': 'Quantity',
            'D1': 'NPS Pipe Size (inches)',
            'E1': 'Working Fluid',
            'F1': 'Process Temp or Pressure',
            'G1': 'System Efficiency (%)',
            'H1': 'Ambient Temp',
            'I1': 'System Hours / Year',
            'J1': 'Wind Speed (MPH)',
            'K1': 'Location',
            'L1': 'Base Metal',
            'M1': 'Insulation',
            'N1': 'Insulation Thickness',
            'O1': 'Jacket Material',
        }
        self.data_column = {
            'A': 'name',
            'B': 'valve_type',
            'C': 'quantity',
            'D': 'nps_pipe_size_inches',
            'E': 'working_fluid',
            'F': 'process_temp_or_pressure',
            'G': 'system_efficiency',
            'H': 'ambient_temp',
            'I': 'system_hours_per_year',
            'J': 'wind_speed_mph',
            'K': 'location',
            'L': 'base_metal',
            'M': 'insulation',
            'N': 'insulation_thickness',
            'O': 'jacket_material',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        for each_valve_insulation in self.valve_insulation:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_valve_insulation.get(column_val)
                # import pdb; pdb.set_trace()
                insulation_choices_dict = dict(INSULATION_CHOICES)
                if column_val in ['insulation']:
                    val = insulation_choices_dict[
                        each_valve_insulation.get(column_val)]
                    self.ws[cell_val] = val
                    self.ws.column_dimensions[each_column].width = len(val)
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'ValveInsulation.xlsx')
        return response


class PipeInsulationExcel(object):
    def __init__(self, pipe_insulation=None):
        self.pipe_insulation = pipe_insulation
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Pipe Insulation Description',
            'B1': 'Length of Pipe (feet)',
            'C1': 'NPS Pipe Size (inches)',
            'D1': 'Working Fluid',
            'E1': 'Process Temp or Pressure',
            'F1': 'System Efficiency (%)',
            'G1': 'Ambient Temp',
            'H1': 'System Hours / Year',
            'I1': 'Wind Speed (MPH)',
            'J1': 'Location',
            'K1': 'Base Metal',
            'L1': 'Insulation',
            'M1': 'Insulation Thickness',
            'N1': 'Jacket Material',
        }
        self.data_column = {
            'A': 'name',
            'B': 'length_of_pipe',
            'C': 'nps_pipe_size_inches',
            'D': 'working_fluid',
            'E': 'process_temp_or_pressure',
            'F': 'system_efficiency',
            'G': 'ambient_temp',
            'H': 'system_hours_per_year',
            'I': 'wind_speed_mph',
            'J': 'location',
            'K': 'base_metal',
            'L': 'insulation',
            'M': 'insulation_thickness',
            'N': 'jacket_material',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        insulation_choices_dict = dict(INSULATION_CHOICES)
        for each_pipe_insulation in self.pipe_insulation:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_pipe_insulation.get(column_val)
                if column_val in ['insulation']:
                    val = insulation_choices_dict[
                        each_pipe_insulation.get(column_val)]
                    self.ws[cell_val] = val
                    self.ws.column_dimensions[each_column].width = len(val)
            count += 1

    def get_excel_raw(self):
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'ValveInsulation.xlsx')
        return response


class PremiumEfficiencyExcel(object):
    def __init__(self, premium_efficiency=None):
        self.premium_efficiency = premium_efficiency
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A2': 'Premium Efficiency Name',
            'B2': 'Annual Operating Hours',
            'C2': 'Motor Nameplate HP',
            'D2': 'Motor Nameplate RPM',
            'E2': 'Full load efficiency (%)',
            'F2': '3/4 load efficiency (%)',
            'G2': '1/2 load efficiency (%)',
            'H2': 'Energy cost @ Full Load(annual operating cost)',
            'I2': 'Energy cost @ 3/4(annual operating cost)',
            'J2': 'Energy cost @ 1/2(annual operating cost)',
            'K2': 'Motor Purchase Price ($)',
            'L2': 'Full load efficiency (%)',
            'M2': '3/4 load efficiency (%)',
            'N2': '1/2 load efficiency (%)',
            'O2': 'Energy cost @ Full Load(annual operating cost)',
            'P2': 'Energy cost @ 3/4(annual operating cost)',
            'Q2': 'Energy cost @ 1/2(annual operating cost)',
            'R2': 'Motor Purchase Price ($)',
            'S2': 'Purchase Price Differential',
            'T2': 'Annual Operating Cost Differiential @ Full Load',
            'U2': 'Annual Operating Cost Differiential @ 3/4',
            'V2': 'Annual Operating Cost Differiential @ 1/2'
        }
        self.data_column = {
            'A': 'motor_name',
            'B': 'annual_operating_hours',
            'C': 'motor_nameplate_hp',
            'D': 'motor_nameplate_rpm',
            'E': 'existing_full_load_eff',
            'F': 'existing_three_fourth_load_eff',
            'G': 'existing_half_load_eff',
            'H': 'get_existing_energy_cost_full_load',
            'I': 'get_existing_energy_cost_three_fourth_load',
            'J': 'get_existing_energy_cost_half_load',
            'K': 'existing_motor_purchase_price',
            'L': 'proposed_full_load_eff',
            'M': 'proposed_three_fourth_load_eff',
            'N': 'proposed_half_load_eff',
            'O': 'get_proposed_energy_cost_full_load',
            'P': 'get_proposed_energy_cost_three_fourth_load',
            'Q': 'get_proposed_energy_cost_half_load',
            'R': 'proposed_motor_purchase_price',
            'S': 'get_purchase_price_diff',
            'T': 'get_energy_cost_full_load_diff',
            'U': 'get_energy_cost_three_fourth_load_diff',
            'V': 'get_energy_cost_half_load_diff',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        # Cell Properties
        cell_existing = {'alignment': Alignment(horizontal='center'),
                         'font': Font(bold=True),
                         'fill': PatternFill(patternType='solid',
                                             fgColor=Color('90909000'))}
        self.ws['E1'] = "Existing Motor"
        self.ws['E1'].style = Style(**cell_existing)
        # Mergeing Cells
        self.ws.merge_cells('E1:K1')
        # Cell Properties
        cell_proposed = {'alignment': Alignment(horizontal='center'),
                         'font': Font(bold=True),
                         'fill': PatternFill(patternType='solid',
                                             fgColor=Color('90909099'))}
        self.ws['L1'] = "Proposed Premium Efficiency Motor"
        self.ws['L1'].style = Style(**cell_proposed)
        # Mergeing Cells
        self.ws.merge_cells('L1:R1')
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('2', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 3
        for each_premium_efficiency in self.premium_efficiency:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_premium_efficiency.get(column_val)

            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'PremiumEfficiency.xlsx')
        return response


class BoilerBlowdownExcel(object):
    def __init__(self, boiler_blowdown=None):
        self.boiler_blowdown = boiler_blowdown
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Boiler Blowdown Reduction and Recovery',
            'B1': 'Makeup Water Temperature',
            'C1': 'Blowdown Frequency (Per Day)',
            'D1': 'Blowdown Rate (GPM)',
            'E1': 'Blowdown Duration (minutes)',
            'F1': 'Blowdown (Gallons Per Day)',
            'G1': 'Days of Operation',
            'H1': 'Annual Quantity (Gals)',
            'I1': 'Annual Quantity (lbs)',
            'J1': 'Discharge Temp (F)',
            'K1': 'Blowdown Energy Loss (BTUH)',
            'L1': 'Blowdown Energy Loss (THERM)',
            'M1': 'Blowdown Energy Cost ($)',
            'N1': 'Make-up Water Quantity',
            'O1': 'Make-up Water Cost ($)',
            'P1': 'Gas and Water Cost ($)',
            'Q1': 'Blowdown Energy Loss (Therms)',
            'R1': 'Heat Recovery Efficiency',
            'S1': 'Blowdown Energy Recovery (Therms)',
            'T1': 'Total Savings ($)',
        }
        self.data_column_existing = {
            'A': 'boiler_blowdown_name',
            'B': 'makeup_water_temperature',
            'C': 'existing_blowdown_frequency_per_day',
            'D': 'existing_blowdown_rate_gpm',
            'E': 'existing_blowdown_duration_mins',
            'F': 'get_existing_blowdown_gallons_per_day',
            'G': 'existing_days_of_operation',
            'H': 'get_existing_annual_quantity_gals',
            'I': 'get_existing_annual_quantity_lbs',
            'J': 'existing_discharge_temp_in_f',
            'K': 'get_existing_blowdown_energy_loss_btuh',
            'L': 'get_existing_energy_loss_therm',
            'M': 'get_existing_overflow_cost',
            'N': 'get_existing_makeup_water_quantity',
            'O': 'get_existing_makeup_water_cost',
            'P': 'get_existing_gas_and_water_cost',
            'Q': 'get_existing_blowdown_energy_loss_therm',
            'R': 'existing_heat_recovery_efficiency_perc',
            'S': 'get_existing_blowdown_energy_recovery_therms',
            'T': 'get_existing_total'
        }
        self.data_column_proposed = {
            'A': 'boiler_blowdown_name',
            'B': 'makeup_water_temperature',
            'C': 'proposed_blowdown_frequency_per_day',
            'D': 'proposed_blowdown_rate_gpm',
            'E': 'proposed_blowdown_duration_mins',
            'F': 'get_proposed_blowdown_gallons_per_day',
            'G': 'proposed_days_of_operation',
            'H': 'get_proposed_annual_quantity_gals',
            'I': 'get_proposed_annual_quantity_lbs',
            'J': 'proposed_discharge_temp_in_f',
            'K': 'get_proposed_blowdown_energy_loss_btuh',
            'L': 'get_proposed_energy_loss_therm',
            'M': 'get_proposed_overflow_cost',
            'N': 'get_proposed_makeup_water_quantity',
            'O': 'get_proposed_makeup_water_cost',
            'P': 'get_proposed_gas_and_water_cost',
            'Q': 'get_proposed_blowdown_energy_loss_therm',
            'R': 'proposed_heat_recovery_efficiency_perc',
            'S': 'get_proposed_blowdown_energy_recovery_therms',
            'T': 'get_proposed_total'
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        cell_proposed = {'alignment': Alignment(horizontal='right'),
                         'font': Font(bold=True),
                         'fill': PatternFill(patternType='solid',
                                             fgColor=Color('90909000'))}
        for each_boiler_blowdown in self.boiler_blowdown:
            for each_column in self.data_column_existing:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column_existing[each_column]
                self.ws[cell_val] = each_boiler_blowdown.get(column_val)
            count += 1
            for each_column in self.data_column_proposed:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column_proposed[each_column]
                self.ws[cell_val] = each_boiler_blowdown.get(column_val)
            self.ws['A%s' % (count,)] = "Proposed Blowdown"
            self.ws['A%s' % (count,)].style = Style(**cell_proposed)
            self.ws.merge_cells('A%s:B%s' % (count, count))
            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'BoilerBlowdown.xlsx')
        return response


class AirCompressorExcel(object):
    def __init__(self, air_compressors=None):
        self.air_compressors = air_compressors
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Compressor Name',
            'B1': 'Customer Name',
            'C1': 'Project Name',
            'D1': 'Manufacturer',
            'E1': 'Model #',
            'F1': 'Serial #',
            'G1': 'Type',
            'H1': 'VFD Speed Control',
            'I1': 'Nameplate Horsepower (HP)',
            'J1': 'Nameplate Maximum Flow (CFM)',
            'K1': 'Measured Actual Flow (GPM)',
            'L1': 'Measured Line Pressure (PSI)',
            'M1': 'Annual Hours Of Operation (Hours)',
            'N1': 'Hourly kWh Consumed (kWh)',
            'O1': 'Hourly Cost Of Operation ($)',
            'P1': 'Annual Cost Of Operation ($)',
            'Q1': 'Reduced Line Pressure TO (PSI)',
            'R1': 'Proposed Pressure Decrease (PSI)',
            'S1': 'Estimated Annual Savings Per 2 PSI Reduction ($)',
            'T1': 'Annual Cost Before PSI Setback ($)',
            'U1': 'Annual Cost After PSI Setback ($)',
            'V1': 'Annual Savings After PSI Setback ($)'
        }
        self.data_column = {
            'A': 'compressor_name',
            'B': 'customer_name',
            'C': 'project_name',
            'D': 'manufacturer',
            'E': 'model_info',
            'F': 'serial_info',
            'G': 'compressor_type',
            'H': 'vfd_speed_control',
            'I': 'nameplate_horsepower',
            'J': 'nameplate_max_flow',
            'K': 'measured_actual_flow',
            'L': 'measured_line_pressure',
            'M': 'annual_hours_of_operation',
            'N': 'get_hourly_kwh_consumed',
            'O': 'get_hourly_cost_of_operation',
            'P': 'get_annual_cost_of_operation',
            'Q': 'reduce_line_pressure_to',
            'R': 'get_proposed_pressure_decrease',
            'S': 'get_estimated_ann_savings_per_2_psi_reduction',
            'T': 'get_annual_cost_before_psi_setback',
            'U': 'get_annual_cost_after_psi_setback',
            'V': 'get_annual_savings_after_psi_setback',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        # Mergeing Cells
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        yes_no_dict = dict(VFD_CONTROL_TYPE)
        compressor_type_dict = dict(COMPRESSOR_TYPE)
        for each_air_compressors in self.air_compressors:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_air_compressors.get(column_val)
                if column_val in ['vfd_speed_control']:
                    val = yes_no_dict[each_air_compressors.get(column_val)]
                    self.ws[cell_val] = val
                if column_val in ['compressor_type']:
                    val = compressor_type_dict[
                        each_air_compressors.get(column_val)]
                    self.ws[cell_val] = val
                    self.ws.column_dimensions[each_column].width = len(val)

            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'AirCompressor.xlsx')
        return response


class AirLeakExcel(object):
    def __init__(self, air_leak=None):
        self.air_leak = air_leak
        self.wb = Workbook()
        self.ws = self.wb.active
        self.out = StringIO.StringIO()
        self.headers_dict = {
            'A1': 'Air Leak',
            'B1': 'Project Name',
            'C1': 'Leak Tag Number',
            'D1': 'Date Leak found',
            'E1': 'Leak Area Description',
            'F1': 'Leak Equipment Description',
            'G1': 'Leak Type',
            'H1': 'Annual Hours of Operation (Hours)',
            'I1': 'Leak DB Reading',
            'J1': 'Leak Repaired?',
            'K1': 'Convert DB to CFM',
            'L1': 'Annual Cost Of Leak'
        }
        self.data_column = {
            'A': 'air_leak',
            'B': 'project_name',
            'C': 'leak_tag_number',
            'D': 'datetime_time_leak_found',
            'E': 'leak_area_description',
            'F': 'leak_equipment_desc',
            'G': 'leak_type',
            'H': 'annual_hours_of_operation',
            'I': 'leak_db_reading',
            'J': 'leak_reparied_flag',
            'K': 'get_convert_db_to_cmf',
            'L': 'get_annual_cost_of_leak',
        }
        self.make_excel_headers()
        self.make_excel_data()

    def make_excel_headers(self):
        cell_bold = {'font': Font(bold=True)}
        # Mergeing Cells
        for each_cell in self.headers_dict:
            self.ws[each_cell] = self.headers_dict[each_cell]
            self.ws[each_cell].style = Style(**cell_bold)
            column = each_cell.replace('1', '')
            self.ws.column_dimensions[column].width = len(
                self.headers_dict[each_cell])

    def make_excel_data(self):
        count = 2
        yes_no_dict = dict(VFD_CONTROL_TYPE)
        for each_air_leak in self.air_leak:
            for each_column in self.data_column:
                cell_val = '%s%s' % (each_column, count)
                column_val = self.data_column[each_column]
                self.ws[cell_val] = each_air_leak.get(column_val)
                if column_val in ['leak_reparied_flag']:
                    val = yes_no_dict[
                        each_air_leak.get(column_val)]
                    self.ws[cell_val] = val
                    # self.ws.column_dimensions[each_column].width = len(val)

            count += 1

    def get_excel_raw(self):
        # self.wb.read_only = True
        self.wb.save(self.out)
        self.out.seek(0)
        response = HttpResponse(self.out,
                                content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = ('attachment; filename="%s"'
                                           % 'AirLeak.xlsx')
        return response
